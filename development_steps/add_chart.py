from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
workbook = load_workbook('automatedpivot_table.xlsx')
sheet=workbook['Report']

min_column=workbook.active.min_column
max_column=workbook.active.max_column
min_row=workbook.active.min_row
max_row=workbook.active.max_row

barchart=BarChart()
data=Reference(sheet,min_col=min_column+1,max_col=max_column,min_row=min_row,max_row=max_row)
categories=Reference(sheet,min_col=min_column,max_col=min_column,min_row=min_row+1,max_row=max_row)

barchart.add_data(data,titles_from_data=True)
barchart.set_categories(categories)
sheet.add_chart(barchart,'B15')
barchart.title="Sales by Product line"
barchart.style=5
workbook.save('barchart.xlsx')